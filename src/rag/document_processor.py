"""
Document processing and chunking utilities.
"""

import os
from typing import List, Dict, Any
from pathlib import Path
import hashlib
from ..utils.logger import get_logger
from ..config import config

logger = get_logger(__name__)


class DocumentProcessor:
    """Process and chunk documents for RAG"""
    
    def __init__(self, chunk_size: int = None, chunk_overlap: int = None):
        self.chunk_size = chunk_size or config.chunk_size
        self.chunk_overlap = chunk_overlap or config.chunk_overlap
    
    def process_text_file(self, file_path: str) -> List[Dict[str, Any]]:
        """Read and process a text file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            chunks = self.chunk_text(content)
            file_name = os.path.basename(file_path)
            
            documents = []
            for i, chunk in enumerate(chunks):
                doc_id = self._generate_doc_id(file_name, i, chunk)
                documents.append({
                    "id": doc_id,
                    "content": chunk,
                    "metadata": {
                        "source": file_name,
                        "chunk_index": i,
                        "total_chunks": len(chunks)
                    }
                })
            
            logger.info(f"Processed {file_name} into {len(documents)} chunks")
            return documents
        except Exception as e:
            logger.error(f"Error processing text file: {e}")
            return []
    
    def process_pdf_file(self, file_path: str) -> List[Dict[str, Any]]:
        """Process a PDF file (requires pypdf)"""
        try:
            from pypdf import PdfReader
            
            reader = PdfReader(file_path)
            content = ""
            
            for page_num, page in enumerate(reader.pages):
                try:
                    extracted_text = page.extract_text()
                    if extracted_text:
                        content += f"\n--- Page {page_num + 1} ---\n"
                        content += extracted_text
                except Exception as e:
                    logger.warning(f"Error extracting text from page {page_num}: {e}")
            
            if not content.strip():
                logger.warning(f"No text extracted from PDF: {file_path}")
                return []
            
            chunks = self.chunk_text(content)
            file_name = os.path.basename(file_path)
            
            documents = []
            for i, chunk in enumerate(chunks):
                if chunk.strip():  # Only add non-empty chunks
                    doc_id = self._generate_doc_id(file_name, i, chunk)
                    documents.append({
                        "id": doc_id,
                        "content": chunk,
                        "metadata": {
                            "source": file_name,
                            "chunk_index": i,
                            "total_chunks": len(chunks),
                            "file_type": "pdf"
                        }
                    })
            
            logger.info(f"Processed PDF {file_name} into {len(documents)} chunks")
            return documents
        except ImportError:
            logger.error("pypdf not installed. Install with: pip install -r requirements.txt")
            return []
        except Exception as e:
            logger.error(f"Error processing PDF file: {e}")
            return []
    
    def process_markdown_file(self, file_path: str) -> List[Dict[str, Any]]:
        """Process a markdown file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # Split by headers for better structure preservation
            chunks = self.chunk_markdown(content)
            file_name = os.path.basename(file_path)
            
            documents = []
            for i, chunk in enumerate(chunks):
                doc_id = self._generate_doc_id(file_name, i, chunk)
                documents.append({
                    "id": doc_id,
                    "content": chunk,
                    "metadata": {
                        "source": file_name,
                        "chunk_index": i,
                        "total_chunks": len(chunks),
                        "file_type": "markdown"
                    }
                })
            
            logger.info(f"Processed {file_name} into {len(documents)} chunks")
            return documents
        except Exception as e:
            logger.error(f"Error processing markdown file: {e}")
            return []
    
    def process_csv_file(self, file_path: str) -> List[Dict[str, Any]]:
        """Process a CSV file"""
        try:
            import csv
            
            file_name = os.path.basename(file_path)
            rows = []
            
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                if reader.fieldnames is None:
                    logger.warning(f"CSV file {file_name} is empty or has no headers")
                    return []
                
                for row_num, row in enumerate(reader, start=2):  # start=2 to account for header row
                    # Convert row to readable text format
                    row_text = " | ".join([f"{k}: {v}" for k, v in row.items() if v])
                    rows.append(row_text)
            
            if not rows:
                logger.warning(f"No data found in CSV file: {file_name}")
                return []
            
            # Combine rows with header info
            header_row = " | ".join(reader.fieldnames) if reader.fieldnames else "CSV Data"
            content = f"CSV Headers: {header_row}\n\n" + "\n".join(rows)
            
            chunks = self.chunk_text(content)
            documents = []
            
            for i, chunk in enumerate(chunks):
                doc_id = self._generate_doc_id(file_name, i, chunk)
                documents.append({
                    "id": doc_id,
                    "content": chunk,
                    "metadata": {
                        "source": file_name,
                        "chunk_index": i,
                        "total_chunks": len(chunks),
                        "file_type": "csv"
                    }
                })
            
            logger.info(f"Processed CSV {file_name} into {len(documents)} chunks")
            return documents
        except Exception as e:
            logger.error(f"Error processing CSV file: {e}")
            return []
    
    def process_xlsx_file(self, file_path: str) -> List[Dict[str, Any]]:
        """Process an XLSX (Excel) file"""
        try:
            from openpyxl import load_workbook
            
            file_name = os.path.basename(file_path)
            workbook = load_workbook(file_path)
            all_content = []
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                sheet_text = f"Sheet: {sheet_name}\n"
                
                # Get headers
                headers = []
                for cell in worksheet[1]:
                    headers.append(str(cell.value) if cell.value else "")
                
                sheet_text += "Headers: " + " | ".join(headers) + "\n\n"
                
                # Get data rows
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                    row_text = " | ".join([f"{headers[i]}: {str(val)}" if i < len(headers) else str(val) 
                                          for i, val in enumerate(row) if val is not None])
                    if row_text.strip():
                        sheet_text += row_text + "\n"
                
                all_content.append(sheet_text)
            
            if not all_content or not any(c.strip() for c in all_content):
                logger.warning(f"No data found in XLSX file: {file_name}")
                return []
            
            content = "\n".join(all_content)
            chunks = self.chunk_text(content)
            documents = []
            
            for i, chunk in enumerate(chunks):
                doc_id = self._generate_doc_id(file_name, i, chunk)
                documents.append({
                    "id": doc_id,
                    "content": chunk,
                    "metadata": {
                        "source": file_name,
                        "chunk_index": i,
                        "total_chunks": len(chunks),
                        "file_type": "xlsx"
                    }
                })
            
            logger.info(f"Processed XLSX {file_name} into {len(documents)} chunks")
            return documents
        except ImportError:
            logger.error("openpyxl not installed. Install with: pip install -r requirements.txt")
            return []
        except Exception as e:
            logger.error(f"Error processing XLSX file: {e}")
            return []
    
    def process_xls_file(self, file_path: str) -> List[Dict[str, Any]]:
        """Process an XLS (Excel) file"""
        try:
            import xlrd
            
            file_name = os.path.basename(file_path)
            workbook = xlrd.open_workbook(file_path)
            all_content = []
            
            for sheet_idx, worksheet in enumerate(workbook.sheets()):
                sheet_text = f"Sheet: {worksheet.name}\n"
                
                # Get headers
                headers = []
                if worksheet.nrows > 0:
                    for col_idx in range(worksheet.ncols):
                        headers.append(str(worksheet.cell_value(0, col_idx)))
                
                sheet_text += "Headers: " + " | ".join(headers) + "\n\n"
                
                # Get data rows
                for row_idx in range(1, worksheet.nrows):
                    row_values = []
                    for col_idx in range(worksheet.ncols):
                        cell_value = worksheet.cell_value(row_idx, col_idx)
                        if cell_value:
                            row_values.append(f"{headers[col_idx]}: {str(cell_value)}")
                    
                    if row_values:
                        sheet_text += " | ".join(row_values) + "\n"
                
                all_content.append(sheet_text)
            
            if not all_content or not any(c.strip() for c in all_content):
                logger.warning(f"No data found in XLS file: {file_name}")
                return []
            
            content = "\n".join(all_content)
            chunks = self.chunk_text(content)
            documents = []
            
            for i, chunk in enumerate(chunks):
                doc_id = self._generate_doc_id(file_name, i, chunk)
                documents.append({
                    "id": doc_id,
                    "content": chunk,
                    "metadata": {
                        "source": file_name,
                        "chunk_index": i,
                        "total_chunks": len(chunks),
                        "file_type": "xls"
                    }
                })
            
            logger.info(f"Processed XLS {file_name} into {len(documents)} chunks")
            return documents
        except ImportError:
            logger.error("xlrd not installed. Install with: pip install -r requirements.txt")
            return []
        except Exception as e:
            logger.error(f"Error processing XLS file: {e}")
            return []
    
    def process_file(self, file_path: str) -> List[Dict[str, Any]]:
        """Process any supported file type
        
        Supported formats:
        - .txt: Plain text files
        - .pdf: PDF documents
        - .md, .markdown: Markdown files
        - .csv: Comma-separated values
        - .xls: Excel 97-2003 workbook
        - .xlsx: Excel 2007+ workbook
        """
        file_ext = Path(file_path).suffix.lower()
        
        if file_ext == '.pdf':
            return self.process_pdf_file(file_path)
        elif file_ext in ['.md', '.markdown']:
            return self.process_markdown_file(file_path)
        elif file_ext == '.csv':
            return self.process_csv_file(file_path)
        elif file_ext == '.xlsx':
            return self.process_xlsx_file(file_path)
        elif file_ext == '.xls':
            return self.process_xls_file(file_path)
        else:
            # Default to text file processing
            return self.process_text_file(file_path)
    
    def chunk_text(self, text: str) -> List[str]:
        """Split text into overlapping chunks"""
        chunks = []
        step = self.chunk_size - self.chunk_overlap
        
        for i in range(0, len(text), step):
            chunk = text[i:i + self.chunk_size]
            if chunk.strip():
                chunks.append(chunk)
        
        return chunks if chunks else [text]
    
    def chunk_markdown(self, text: str) -> List[str]:
        """Split markdown into chunks by headers when possible"""
        chunks = []
        current_chunk = ""
        
        lines = text.split('\n')
        for line in lines:
            # Check if line is a header
            if line.startswith('#'):
                if current_chunk and len(current_chunk) > self.chunk_size // 2:
                    chunks.append(current_chunk.strip())
                    current_chunk = line + "\n"
                else:
                    current_chunk += line + "\n"
            else:
                current_chunk += line + "\n"
                
                # Split if chunk is too large
                if len(current_chunk) > self.chunk_size:
                    chunks.append(current_chunk.strip())
                    current_chunk = ""
        
        if current_chunk.strip():
            chunks.append(current_chunk.strip())
        
        return chunks if chunks else [text]
    
    @staticmethod
    def _generate_doc_id(filename: str, chunk_idx: int, content: str) -> str:
        """Generate a unique document ID"""
        content_hash = hashlib.md5(content.encode()).hexdigest()[:8]
        return f"{filename}_{chunk_idx}_{content_hash}"
